from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
import os
import io

app = Flask(__name__)

# Name of your template file (ensure it's in the same folder)
EXCEL_FILENAME = "MTV-QC-FM-013A_Rev.00 - MTC.xlsx"
EXCEL_PATH = os.path.join(os.getcwd(), EXCEL_FILENAME)

@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    data = request.get_json()
    print("Received data:", data)

    # Load Excel template
    workbook = load_workbook(EXCEL_PATH)
    
    # Use the correct sheet name as visible in Excel
    if "Page 01 of 02" in workbook.sheetnames:
        sheet = workbook["Page 01 of 02"]
    else:
        sheet = workbook.active  # fallback

    def safe_write(cell, value):
        try:
            sheet[cell] = value
        except Exception as e:
            print(f"Error writing to {cell}: {e}")

    # Write values to specific cells
    safe_write('F9', data.get('CUSTOMER_NAME', ''))
    safe_write('F10', data.get('CUSTOMER_PURCHASE_ORDER_NUMBER', ''))
    safe_write('F12', data.get('MTV_ORDER_NUMBER', ''))
    safe_write('F13', data.get('MTV_ORDER_ITEM_NUMBER', ''))
    safe_write('C15', data.get('TYPE', ''))
    safe_write('C16', data.get('SIZE', ''))
    safe_write('C17', data.get('CLASS', ''))
    safe_write('C21', data.get('CONFIGURATION', ''))
    safe_write('C20', data.get('OPERATOR', ''))
    safe_write('C24', data.get('ACCEPTED_QUANTITY', ''))

    # Save to memory
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Save also to disk
    with open("latest.xlsx", "wb") as f:
        f.write(output.getbuffer())

    return jsonify({"url": f"https://{request.host}/download"})

@app.route('/download')
def download_file():
    return send_file("latest.xlsx", as_attachment=True, download_name="MTC-Generated.xlsx")

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
